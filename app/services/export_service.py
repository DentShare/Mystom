"""Экспорт базы пациентов в Excel: пациенты, записи на приём, история лечения, импланты."""
from datetime import datetime, date
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database.models import (
    Patient,
    Appointment,
    Treatment,
    ImplantLog,
)


def _date_fmt(d: datetime | date | None) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d.%m.%Y %H:%M")
    return d.strftime("%d.%m.%Y")


async def get_patients_with_relations(
    db: AsyncSession,
    doctor_id: int,
) -> List[Patient]:
    """Все пациенты врача с записями, лечением и имплантами."""
    stmt = (
        select(Patient)
        .where(Patient.doctor_id == doctor_id)
        .options(
            selectinload(Patient.appointments).selectinload(Appointment.service),
            selectinload(Patient.appointments).selectinload(Appointment.location),
            selectinload(Patient.treatments),
            selectinload(Patient.implant_logs),
        )
        .order_by(Patient.full_name)
    )
    result = await db.execute(stmt)
    return list(result.scalars().unique().all())


def build_patients_excel(patients: List[Patient]) -> BytesIO:
    """Собрать Excel: листы Пациенты, Записи, История лечения, Импланты."""
    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Лист "Пациенты" ---
    ws_p = wb.active
    ws_p.title = "Пациенты"
    headers_p = ["ID", "ФИО", "Телефон", "Дата рождения", "Заметки", "Создан"]
    for col, h in enumerate(headers_p, 1):
        ws_p.cell(row=1, column=col, value=h, font=bold)
    for row_idx, p in enumerate(patients, 2):
        ws_p.cell(row=row_idx, column=1, value=p.id)
        ws_p.cell(row=row_idx, column=2, value=p.full_name or "")
        ws_p.cell(row=row_idx, column=3, value=p.phone or "")
        ws_p.cell(row=row_idx, column=4, value=_date_fmt(p.birth_date) if p.birth_date else "")
        ws_p.cell(row=row_idx, column=5, value=(p.notes or "")[:5000])
        ws_p.cell(row=row_idx, column=5).alignment = wrap
        ws_p.cell(row=row_idx, column=6, value=_date_fmt(p.created_at))

    # --- Лист "Записи на приём" ---
    ws_a = wb.create_sheet("Записи на приём")
    headers_a = ["ID пациента", "ФИО пациента", "Дата и время", "Услуга/описание", "Длительность (мин)", "Статус", "Локация", "Создан"]
    for col, h in enumerate(headers_a, 1):
        ws_a.cell(row=1, column=col, value=h, font=bold)
    row_idx = 2
    for p in patients:
        for a in sorted(p.appointments, key=lambda x: x.date_time):
            service_text = (a.service.name if a.service else None) or a.service_description or ""
            loc_text = a.location.name if a.location else ""
            ws_a.cell(row=row_idx, column=1, value=p.id)
            ws_a.cell(row=row_idx, column=2, value=p.full_name or "")
            ws_a.cell(row=row_idx, column=3, value=_date_fmt(a.date_time))
            ws_a.cell(row=row_idx, column=4, value=service_text)
            ws_a.cell(row=row_idx, column=5, value=getattr(a, "duration_minutes", None) or "")
            ws_a.cell(row=row_idx, column=6, value=a.status or "")
            ws_a.cell(row=row_idx, column=7, value=loc_text)
            ws_a.cell(row=row_idx, column=8, value=_date_fmt(a.created_at))
            row_idx += 1

    # --- Лист "История лечения" ---
    ws_t = wb.create_sheet("История лечения")
    headers_t = [
        "ID пациента", "ФИО", "Дата", "Услуга", "Комментарий", "Зуб",
        "Цена", "Скидка %", "Скидка сумма", "Оплачено", "Способ оплаты", "Статус оплаты"
    ]
    for col, h in enumerate(headers_t, 1):
        ws_t.cell(row=1, column=col, value=h, font=bold)
    row_idx = 2
    for p in patients:
        for t in sorted(p.treatments, key=lambda x: x.created_at or datetime.min):
            ws_t.cell(row=row_idx, column=1, value=p.id)
            ws_t.cell(row=row_idx, column=2, value=p.full_name or "")
            ws_t.cell(row=row_idx, column=3, value=_date_fmt(t.created_at))
            ws_t.cell(row=row_idx, column=4, value=t.service_name or "")
            ws_t.cell(row=row_idx, column=5, value=(t.treatment_notes or "")[:2000])
            ws_t.cell(row=row_idx, column=5).alignment = wrap
            ws_t.cell(row=row_idx, column=6, value=t.tooth_number or "")
            ws_t.cell(row=row_idx, column=7, value=t.price if t.price is not None else "")
            ws_t.cell(row=row_idx, column=8, value=t.discount_percent if t.discount_percent is not None else "")
            ws_t.cell(row=row_idx, column=9, value=t.discount_amount if t.discount_amount is not None else "")
            ws_t.cell(row=row_idx, column=10, value=t.paid_amount if t.paid_amount is not None else "")
            ws_t.cell(row=row_idx, column=11, value=t.payment_method or "")
            ws_t.cell(row=row_idx, column=12, value=t.payment_status or "")
            row_idx += 1

    # --- Лист "Импланты" ---
    ws_i = wb.create_sheet("Импланты")
    headers_i = ["ID пациента", "ФИО", "Зуб", "Система", "Размер", "Дата операции", "Заметки"]
    for col, h in enumerate(headers_i, 1):
        ws_i.cell(row=1, column=col, value=h, font=bold)
    row_idx = 2
    for p in patients:
        for imp in p.implant_logs:
            ws_i.cell(row=row_idx, column=1, value=p.id)
            ws_i.cell(row=row_idx, column=2, value=p.full_name or "")
            ws_i.cell(row=row_idx, column=3, value=imp.tooth_number or "")
            ws_i.cell(row=row_idx, column=4, value=imp.system_name or "")
            ws_i.cell(row=row_idx, column=5, value=imp.implant_size or "")
            ws_i.cell(row=row_idx, column=6, value=_date_fmt(imp.operation_date) if imp.operation_date else "")
            ws_i.cell(row=row_idx, column=7, value=(imp.notes or "")[:1000])
            ws_i.cell(row=row_idx, column=7).alignment = wrap
            row_idx += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
